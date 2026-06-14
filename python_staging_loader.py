"""Python loader: Excel/CSV -> staging_workorder_flat (kolom typed) -> CALL sp_etl_workorder.

Tidak ada JSON di pipeline — ini jalur tercepat untuk ~14k+ baris.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

if os.name == "nt":
    os.environ.setdefault("SystemRoot", r"C:\Windows")
    os.environ.setdefault("WINDIR", r"C:\Windows")
    os.environ.setdefault("ComSpec", r"C:\Windows\system32\cmd.exe")

import mysql.connector

FLAT_COLUMNS: Tuple[str, ...] = (
    "wo_sc_id", "sc_id", "track_id", "track_id_baru",
    "tanggal", "tanggal_order", "tanggal_komitmen", "tgl_input_hd_gdocs",
    "sto", "status_wo", "status_sc",
    "kendala_pt1", "kategori_roc", "kategori_solusi", "solusi_kendala",
    "nik_teknisi", "korlap", "komandan_team", "mitra", "spv", "cp",
    "nama_pelanggan", "nama_contact", "segment", "layanan", "alamat_instalasi", "uic", "koordinat_pelanggan",
    "odp", "odc", "gpon", "feeder", "distribusi", "datek1", "datek_inputan", "datek_real",
    "hasil_ukur_odp", "hasil_ukur_distribusi", "hasil_ukur_feeder",
    "durasi_hari", "durasi", "durasi_manja", "durasi_grup", "durasi_grup_pengerjaan",
    "keterangan", "keterangan_sm_provisioning", "keterangan_tl_provisioning",
    "hasil_solusi_maintenance", "hasil_solusi_optima", "hasil_solusi_sdi",
    "total_eskalasi", "jumlah_kendala", "is_unsc",
)

DATE_FIELDS = {"tanggal", "tanggal_order", "tanggal_komitmen", "tgl_input_hd_gdocs"}
DECIMAL_FIELDS = {"durasi_hari", "durasi", "durasi_manja", "durasi_grup_pengerjaan"}
INT_FIELDS = {"total_eskalasi", "jumlah_kendala"}

INSERT_SQL = (
    f"INSERT INTO staging_workorder_flat ({', '.join(FLAT_COLUMNS)}, created_at, updated_at) "
    f"VALUES ({', '.join(['%s'] * len(FLAT_COLUMNS))}, NOW(), NOW())"
)


def _load_env_file(env_path: str) -> None:
    if not os.path.exists(env_path):
        return
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k, v = k.strip(), v.strip().strip('"').strip("'")
            if k and k not in os.environ:
                os.environ[k] = v


def _load_column_mapping(path: Optional[str]) -> Dict[int, str]:
    if not path:
        return {}
    import json
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    columns = payload.get("columns", payload)

    # FIX: Handle jika columns berupa list
    if isinstance(columns, list):
        return {i: str(v) for i, v in enumerate(columns) if v}

    # Handle jika columns berupa dict
    if isinstance(columns, dict):
        return {int(k): str(v) for k, v in columns.items() if v}

    raise ValueError(
        f"Format mapping tidak valid: 'columns' harus berupa dict atau list, "
        f"bukan {type(columns).__name__}. Cek file: {path}"
    )


def _clean_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _parse_date(val: Any) -> Optional[str]:
    if val is None or val == "":
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        serial = int(val)
        if 32874 < serial < 73050:
            return (date(1899, 12, 30) + timedelta(days=serial)).strftime("%Y-%m-%d")
    text = _clean_str(val)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_decimal(val: Any) -> Optional[Decimal]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    text = re.sub(r"[^\d.\-]", "", _clean_str(val).replace(",", "."))
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_int(val: Any, default: Optional[int] = None) -> Optional[int]:
    if val is None or val == "":
        return default
    if isinstance(val, int):
        return val
    text = re.sub(r"[^\d\-]", "", _clean_str(val))
    if not text:
        return default
    try:
        return int(text)
    except ValueError:
        return default


def _parse_flag(val: Any) -> int:
    text = _clean_str(val).lower()
    return 1 if text in {"1", "yes", "ya", "true", "y"} else 0


def _map_row(values: Sequence[Any], column_mapping: Dict[int, str]) -> Optional[Dict[str, Any]]:
    raw: Dict[str, Any] = {}
    has_value = False
    for idx, field in column_mapping.items():
        if idx >= len(values):
            continue
        val = values[idx]
        text = _clean_str(val)
        raw[field] = text
        if text:
            has_value = True
    if not has_value:
        return None

    wo_sc_id = raw.get("wo_sc_id") or None
    track_id = raw.get("track_id") or wo_sc_id
    status_wo = raw.get("status_wo") or raw.get("status") or "UNKNOWN"

    row: Dict[str, Any] = {col: None for col in FLAT_COLUMNS}
    for field, val in raw.items():
        if field in row:
            row[field] = val or None
        if field == "status" and not row.get("status_wo"):
            row["status_wo"] = val or "UNKNOWN"
        if field == "tanggal_komitmen_ps_completed":
            row["tanggal_komitmen"] = _parse_date(val)
        if field in {"unsc", "is_unsc"}:
            row["is_unsc"] = _parse_flag(val)

    row["wo_sc_id"] = wo_sc_id
    row["track_id"] = track_id
    row["status_wo"] = status_wo
    row["sto"] = row.get("sto") or "UNKNOWN"
    row["kendala_pt1"] = row.get("kendala_pt1") or "UNKNOWN"
    row["nik_teknisi"] = row.get("nik_teknisi") or "UNKNOWN"
    row["is_unsc"] = row.get("is_unsc", 0)
    row["jumlah_kendala"] = _parse_int(row.get("jumlah_kendala"), 1)

    for field in DATE_FIELDS:
        if field in raw or field == "tanggal_komitmen":
            parsed = _parse_date(row.get(field) if field != "tanggal_komitmen" else raw.get("tanggal_komitmen_ps_completed") or row.get("tanggal_komitmen"))
            if parsed:
                row[field] = parsed
    for field in DECIMAL_FIELDS:
        if row.get(field) is not None:
            row[field] = _parse_decimal(row.get(field))
    if row.get("total_eskalasi") is not None:
        row["total_eskalasi"] = _parse_int(row.get("total_eskalasi"))

    if not wo_sc_id:
        return None
    return row


def _row_to_tuple(row: Dict[str, Any]) -> Tuple[Any, ...]:
    return tuple(row.get(col) for col in FLAT_COLUMNS)


def _stream_calamine_xlsx(path: str, column_mapping: Dict[int, str]) -> Iterable[Tuple[Any, ...]]:
    from python_calamine import CalamineWorkbook

    sheet = CalamineWorkbook.from_path(path).get_sheet_by_index(0)
    rows = sheet.to_python(skip_empty_area=False)
    for values in rows[1:]:
        if not values:
            continue
        record = _map_row(values, column_mapping)
        if record:
            yield _row_to_tuple(record)


def _stream_openpyxl_xlsx(path: str, column_mapping: Dict[int, str]) -> Iterable[Tuple[Any, ...]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        rows_iter = wb.active.iter_rows(values_only=True)
        next(rows_iter, None)
        for row in rows_iter:
            if row is None:
                continue
            record = _map_row(row, column_mapping)
            if record:
                yield _row_to_tuple(record)
    finally:
        wb.close()


def _stream_xlsx(path: str, column_mapping: Dict[int, str]) -> Iterable[Tuple[Any, ...]]:
    try:
        yield from _stream_calamine_xlsx(path, column_mapping)
    except ImportError:
        print("python-calamine not installed, using openpyxl (slower). pip install python-calamine", flush=True)
        yield from _stream_openpyxl_xlsx(path, column_mapping)


def _stream_csv(path: str, column_mapping: Dict[int, str]) -> Iterable[Tuple[Any, ...]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            record = _map_row(row, column_mapping)
            if record:
                yield _row_to_tuple(record)


def _stream_mapped_file(path: str, column_mapping: Dict[int, str]) -> Iterable[Tuple[Any, ...]]:
    ext = os.path.splitext(path)[1].lower()
    if not column_mapping:
        raise ValueError("--mapping-file wajib untuk file Excel/CSV")
    if ext == ".csv":
        yield from _stream_csv(path, column_mapping)
    elif ext == ".xlsx":
        yield from _stream_xlsx(path, column_mapping)
    elif ext == ".xls":
        import pandas as pd
        df = pd.read_excel(path, dtype=str).fillna("")
        for _, row in df.iterrows():
            record = _map_row(list(row.values), column_mapping)
            if record:
                yield _row_to_tuple(record)
    else:
        raise ValueError(f"Unsupported file extension: {ext}")


def _configure_session(cur) -> None:
    cur.execute("SET SESSION foreign_key_checks = 0")
    cur.execute("SET SESSION unique_checks = 0")
    cur.execute("SET SESSION sql_log_bin = 0")


def _restore_session(cur) -> None:
    cur.execute("SET SESSION foreign_key_checks = 1")
    cur.execute("SET SESSION unique_checks = 1")


def _flush_batch(cur, batch: Sequence[Tuple[Any, ...]]) -> int:
    if not batch:
        return 0
    cur.executemany(INSERT_SQL, batch)
    return len(batch)


def main() -> None:
    _load_env_file(os.path.join(os.path.dirname(__file__), ".env"))

    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--log-id", required=True, type=int)
    ap.add_argument("--mapping-file", required=True)
    ap.add_argument("--db-host", default=os.getenv("DB_HOST", "127.0.0.1"))
    ap.add_argument("--db-port", default=int(os.getenv("DB_PORT", "3306")))
    ap.add_argument("--db-name", default=os.getenv("DB_DATABASE", ""))
    ap.add_argument("--db-user", default=os.getenv("DB_USERNAME", ""))
    ap.add_argument("--db-pass", default=os.getenv("DB_PASSWORD", ""))
    ap.add_argument("--batch-size", default=2000, type=int)
    args = ap.parse_args()

    if not args.db_name or not args.db_user:
        raise SystemExit("Missing DB config.")

    column_mapping = _load_column_mapping(args.mapping_file)
    t0 = time.perf_counter()

    cnx = mysql.connector.connect(
        host=args.db_host,
        port=args.db_port,
        user=args.db_user,
        password=args.db_pass,
        database=args.db_name,
        use_pure=True,
    )
    cnx.autocommit = False
    cur = cnx.cursor()

    try:
        print("Truncating staging_workorder_flat...", flush=True)
        cur.execute("TRUNCATE TABLE staging_workorder_flat")
        cnx.commit()
        _configure_session(cur)

        total = inserted = 0
        batch: List[Tuple[Any, ...]] = []
        load_t0 = time.perf_counter()

        print(f"Loading: {args.file}", flush=True)
        for row_tuple in _stream_mapped_file(args.file, column_mapping):
            total += 1
            batch.append(row_tuple)
            if len(batch) >= args.batch_size:
                inserted += _flush_batch(cur, batch)
                batch = []

        if batch:
            inserted += _flush_batch(cur, batch)

        cnx.commit()
        _restore_session(cur)
        load_elapsed = time.perf_counter() - load_t0
        print(
            f"Flat staging: {inserted}/{total} rows in {load_elapsed:.1f}s "
            f"({inserted / max(load_elapsed, 0.001):.0f} rows/s)",
            flush=True,
        )

        sp_t0 = time.perf_counter()
        print(f"CALL sp_etl_workorder({args.log_id})...", flush=True)
        cur.execute("CALL sp_etl_workorder(%s)", (args.log_id,))
        cnx.commit()
        sp_elapsed = time.perf_counter() - sp_t0

        print(
            f"DONE log={args.log_id}: staging {load_elapsed:.1f}s + SP {sp_elapsed:.1f}s "
            f"= total {time.perf_counter() - t0:.1f}s",
            flush=True,
        )
    except Exception as exc:
        cnx.rollback()
        print(f"ETL failed: {exc}", file=sys.stderr)
        raise
    finally:
        cur.close()
        cnx.close()


if __name__ == "__main__":
    main()
